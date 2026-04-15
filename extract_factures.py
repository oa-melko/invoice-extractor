"""
Extract invoice data from scanned PDFs using docTR OCR and write to Excel.
"""
import os
import re
from datetime import datetime
from doctr.io import DocumentFile
from doctr.models import ocr_predictor
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# ── Configurable parameters ──────────────────────────────────────────
CONFIDENCE_THRESHOLD = 0.80
INPUT_DIR = "donnee a analyser"
OUTPUT_FILE = "Saisie Achat Linge - RESULTAT.xlsx"

# Supplier display names (order matters for Excel output)
SUPPLIER_ORDER = [
    "Cloro'fil Concept",
    "CTM STYLE",
    "HALBOUT SAS",
    "MULLIEZ-FLORY",
    "POYET-MOTTE",
    "TISSUS GISELE",
]

# Map filename prefixes to supplier names
SUPPLIER_PREFIX_MAP = {
    "Clorofil": "Cloro'fil Concept",
    "CTM": "CTM STYLE",
    "Halbout": "HALBOUT SAS",
    "Mulliez": "MULLIEZ-FLORY",
    "Poyet Motte": "POYET-MOTTE",
    "Tissus Gis": "TISSUS GISELE",
}

LOW_CONF_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def detect_supplier(filename):
    """Detect supplier from PDF filename prefix."""
    for prefix, name in SUPPLIER_PREFIX_MAP.items():
        if filename.startswith(prefix):
            return name
    return None


def ocr_pdf(filepath, model):
    """OCR a PDF file and return the docTR export dict."""
    doc = DocumentFile.from_pdf(filepath)
    result = model(doc)
    return result.export()


def extract_lines(ocr_export):
    """Extract text lines with average confidence from OCR export."""
    lines = []
    for page in ocr_export['pages']:
        for block in page['blocks']:
            for line in block['lines']:
                text = ' '.join(w['value'] for w in line['words'])
                confs = [w['confidence'] for w in line['words']]
                avg = sum(confs) / len(confs) if confs else 0.0
                lines.append({
                    'text': text,
                    'confidence': avg,
                    'words': line['words'],
                })
    return lines


# ── Parsers ──────────────────────────────────────────────────────────

def parse_clorofil(lines):
    """Parse Cloro'fil invoice lines.

    Pattern: internal 6-digit code, then description, then qty, then prices,
    then reference code (all-caps alphanumeric), then optional color line.
    """
    items = []
    date, date_conf = None, 0.0

    for line in lines:
        m = re.search(r'du\s+(\d{2}/\d{2}/\d{4})', line['text'])
        if m:
            try:
                date = datetime.strptime(m.group(1), '%d/%m/%Y')
                date_conf = line['confidence']
                break
            except ValueError:
                continue

    i = 0
    while i < len(lines):
        text = lines[i]['text'].strip()

        if re.match(r'^\d{6}$', text):
            description = ''
            desc_conf = 0.0
            quantity = None
            qty_conf = 0.0
            reference = ''
            ref_conf = 0.0

            j = i + 1
            while j < len(lines):
                t = lines[j]['text'].strip()
                if re.match(r'^\d[\d\s]*$', t.replace(' ', '')):
                    quantity = int(t.replace(' ', ''))
                    qty_conf = lines[j]['confidence']
                    j += 1
                    break
                else:
                    if description:
                        description += ' '
                    description += t
                    desc_conf = (desc_conf + lines[j]['confidence']) / 2 if desc_conf > 0 else lines[j]['confidence']
                    j += 1

            while j < len(lines):
                t = lines[j]['text'].strip()
                # Skip price lines (numbers, commas, dots, spaces, colons)
                if re.match(r'^[\d\s,.:]+$', t):
                    j += 1
                else:
                    break

            if j < len(lines):
                t = lines[j]['text'].strip()
                if re.match(r'^[A-Z0-9]+$', t) and len(t) > 4:
                    reference = t
                    ref_conf = lines[j]['confidence']
                    j += 1

                    if j < len(lines):
                        color = lines[j]['text'].strip()
                        if (len(color) < 30 and
                            not re.match(r'^\d', color) and
                            not re.match(r'^[A-Z0-9]{6,}$', color) and
                            not any(kw in color.lower() for kw in [
                                'conditions', 'commerciales', 'attention',
                                'commentaire', 'order', 'commande', 'women',
                                'carton', 'inscription'])):
                            description += ' ' + color
                            desc_conf = (desc_conf + lines[j]['confidence']) / 2
                            j += 1

            if reference and quantity is not None:
                items.append({
                    'fournisseur': "Cloro'fil Concept",
                    'date': date, 'date_conf': date_conf,
                    'reference': reference, 'ref_conf': ref_conf,
                    'article': description, 'article_conf': desc_conf,
                    'composants': None, 'composants_conf': 1.0,
                    'quantite': quantity, 'quantite_conf': qty_conf,
                })

            i = j
        else:
            i += 1

    return items


def parse_chorus_format(lines, supplier_name):
    """Parse Chorus portal invoice format (used by CTM Style and Poyet Motte).

    Pattern: "TVA% N. ARTICLE_NAME" or "TVA% N- ARTICLE_NAME",
    then "QTY,00 (EA)" quantity line, then price, then "Référence produit : EAN".
    """
    items = []
    date, date_conf = None, 0.0

    for line in lines:
        m = re.search(r'du\s+(\d{2}/\d{2}/\d{4})', line['text'])
        if m:
            try:
                dt = datetime.strptime(m.group(1), '%d/%m/%Y')
                if 'acture' in line['text'] or 'Facture' in line['text']:
                    date = dt
                    date_conf = line['confidence']
                    break
                elif date is None:
                    date = dt
                    date_conf = line['confidence']
            except ValueError:
                continue

    i = 0
    while i < len(lines):
        text = lines[i]['text'].strip()

        # Match "20,00 N. ARTICLE" or "20,00 N- ARTICLE" or "20,00 N-ARTICLE"
        m = re.match(r'^[\d,]+\s+\d+[\.\-]\s*(.+)$', text)
        if m:
            article_name = m.group(1).strip()
            if any(kw in article_name.lower() for kw in ['page', 'total', 'livraison', 'remise', 'charge']):
                i += 1
                continue
            # Strip leading item number/letter prefix (e.g. "1 ROMEO" -> "ROMEO", "F ROMEO" -> "ROMEO")
            article_name = re.sub(r'^[A-Z0-9]\s+', '', article_name)

            article_conf = lines[i]['confidence']
            quantity = None
            qty_conf = 0.0
            reference = ''
            ref_conf = 0.0

            j = i + 1
            # Quantity line: "15,00 (EA)" or "15,00" or "70,00 (EA)"
            if j < len(lines):
                t = lines[j]['text'].strip()
                qty_m = re.match(r'^([\d\s]+)[,.]00\s*(?:\(EA\))?', t)
                if not qty_m:
                    qty_m = re.match(r'^([\d\s]+)[,.]?0*$', t)
                if qty_m:
                    try:
                        quantity = int(qty_m.group(1).replace(' ', '').replace(',', ''))
                        qty_conf = lines[j]['confidence']
                    except ValueError:
                        pass
                    j += 1

            # Look for "Référence produit : XXXXX" in next few lines
            for k in range(j, min(j + 6, len(lines))):
                # Try numeric ref first (CTM, Poyet Motte)
                ref_m = re.search(r'[Rr][eéè][fé]f?[eéè]?rence\s+produit\s*[:;.\-]\s*(\d+)', lines[k]['text'])
                if not ref_m:
                    # Try alphanumeric ref (Tissus Gisele: "DRC2 180X320 AI PC")
                    ref_m = re.search(r'[Rr][eéè][fé]f?[eéè]?rence\s+produit\s*[:;.\-]\s*(.+)$', lines[k]['text'])
                if ref_m:
                    reference = ref_m.group(1).strip().lstrip('- ')
                    ref_conf = lines[k]['confidence']
                    break

            if quantity is not None and reference:
                items.append({
                    'fournisseur': supplier_name,
                    'date': date, 'date_conf': date_conf,
                    'reference': reference, 'ref_conf': ref_conf,
                    'article': article_name, 'article_conf': article_conf,
                    'composants': None, 'composants_conf': 1.0,
                    'quantite': quantity, 'quantite_conf': qty_conf,
                })

        i += 1

    return items


def parse_ctm(lines):
    """Parse CTM Style invoice lines (Chorus format)."""
    return parse_chorus_format(lines, 'CTM STYLE')


def parse_halbout(lines):
    """Parse Halbout SAS invoice lines.

    Pattern: reference code (letters+digits), then "QTY UN",
    then price, then total+TVA, then multi-line description.
    """
    items = []
    date, date_conf = None, 0.0

    for line in lines:
        m = re.search(r'le\s+(\d{2}/\d{2}/\d{4})', line['text'])
        if m:
            try:
                date = datetime.strptime(m.group(1), '%d/%m/%Y')
                date_conf = line['confidence']
                break
            except ValueError:
                continue

    i = 0
    while i < len(lines):
        text = lines[i]['text'].strip()

        if re.match(r'^[A-Z]{2,4}\d{2,}[A-Z0-9]*$', text) and len(text) >= 6:
            reference = text
            ref_conf = lines[i]['confidence']

            j = i + 1
            quantity = None
            qty_conf = 0.0
            description_parts = []
            desc_conf_sum = 0.0
            desc_count = 0

            if j < len(lines):
                qty_m = re.match(r'^(\d+)\s+UN', lines[j]['text'].strip())
                if qty_m:
                    quantity = int(qty_m.group(1))
                    qty_conf = lines[j]['confidence']
                    j += 1
                elif re.match(r'^\d+$', lines[j]['text'].strip()):
                    # qty alone, UN might be on next line
                    potential_qty = int(lines[j]['text'].strip())
                    if j + 1 < len(lines) and lines[j + 1]['text'].strip() == 'UN':
                        quantity = potential_qty
                        qty_conf = lines[j]['confidence']
                        j += 2  # skip qty + UN
                    else:
                        i += 1
                        continue
                else:
                    i += 1
                    continue

            while j < len(lines):
                t = lines[j]['text'].strip()
                if re.match(r'^[\d\s.,]+$', t):
                    j += 1
                else:
                    break

            while j < len(lines):
                t = lines[j]['text'].strip()
                if re.match(r'^[A-Z]{2,4}\d{2,}[A-Z0-9]*$', t) and len(t) >= 6:
                    break
                if any(kw in t.lower() for kw in ['livre', 'page ', 'n/ bl', 'n/ ar',
                        'liberatoire', 'reglement', 'taux tva', 'total', 'aucun',
                        'penalite', 'vendeur', 'chorus', 'indemnite', 'pour etre',
                        'directement', 'bnp paribas', 'immeuble', 'dunkerque',
                        'marseille', 'compte', 'subrogation', 'affacturage']):
                    break
                if re.match(r'^[\d\s.,]+$', t):
                    break
                description_parts.append(t)
                desc_conf_sum += lines[j]['confidence']
                desc_count += 1
                j += 1

            description = ' '.join(description_parts)
            desc_conf = desc_conf_sum / desc_count if desc_count > 0 else 0.0

            if quantity is not None and description:
                items.append({
                    'fournisseur': 'HALBOUT SAS',
                    'date': date, 'date_conf': date_conf,
                    'reference': reference, 'ref_conf': ref_conf,
                    'article': description, 'article_conf': desc_conf,
                    'composants': None, 'composants_conf': 1.0,
                    'quantite': quantity, 'quantite_conf': qty_conf,
                })

            i = j
        else:
            i += 1

    return items


def parse_mulliez(lines):
    """Parse Mulliez-Flory invoice lines.

    Two formats:
    - Format A: "REF COLORIS", TVA, description, GENCOD, "TU QTY PRICE"
    - Format B: "REF COLORIS", TVA, description, then repeating
      [GENCOD, size, "QTY PRICE"] blocks (one per size)
    """
    items = []
    date, date_conf = None, 0.0

    for line in lines:
        m = re.search(r'DU\s+(\d{2}/\d{2}/\d{4})', line['text'])
        if m:
            try:
                date = datetime.strptime(m.group(1), '%d/%m/%Y')
                date_conf = line['confidence']
                break
            except ValueError:
                continue

    i = 0
    while i < len(lines):
        text = lines[i]['text'].strip()

        # Article ref line: 6-digit number + space + coloris code (2-6 alphanumeric)
        m = re.match(r'^(\d{6})\s+([A-Z][A-Z0-9]{1,5})$', text)

        if m:
            ref_code = m.group(1)
            coloris = m.group(2)
            reference = f'{ref_code} {coloris}'
            ref_conf = lines[i]['confidence']

            j = i + 1
            desc_conf = 0.0

            # Skip TVA line(s) and noise
            while j < len(lines):
                t = lines[j]['text'].strip()
                if ('TVA' in t.upper() or 'TV�' in t or
                    re.match(r'^[\-\s]*\d{1,2}(\s+\d{1,2})?\s*(TVA|TV)?', t, re.IGNORECASE) or
                    t in ('-', 'DUDE', 'N') or
                    re.match(r'^[\-\s]+\d\s', t)):
                    j += 1
                else:
                    break

            # Collect description lines until GENCOD or TU or next article
            desc_parts = []
            while j < len(lines):
                t = lines[j]['text'].strip()
                if re.match(r'^\d{13}$', t):
                    break  # GENCOD found - don't skip yet
                if t.upper().startswith('TU '):
                    break
                if re.match(r'^\d{6}\s+[A-Z]', t):
                    break
                if t.upper().startswith('TOTAL'):
                    break
                if not re.match(r'^[\d\s,.]+$', t) and len(t) > 2:
                    desc_parts.append(t)
                    desc_conf = (desc_conf + lines[j]['confidence']) / 2 if desc_conf > 0 else lines[j]['confidence']
                j += 1

            description = ' '.join(desc_parts).rstrip('.')
            total_quantity = 0
            qty_conf = 0.0
            gencod_pos = j  # save position for Format B

            # Skip GENCOD line if present (13 digits)
            if j < len(lines) and re.match(r'^\d{13}$', lines[j]['text'].strip()):
                j += 1

            # Format A: "TU QTY PRICE" or "TU" alone then "QTY PRICE" on next line
            if j < len(lines):
                t = lines[j]['text'].strip()
                tu_m = re.match(r'^TU\s+(\d+)\s+', t)
                if tu_m:
                    total_quantity = int(tu_m.group(1))
                    qty_conf = lines[j]['confidence']
                    j += 1
                elif t == 'TU':
                    j += 1
                    if j < len(lines):
                        qty_price_m = re.match(r'^(\d+)\s+[\d,]+', lines[j]['text'].strip())
                        if qty_price_m:
                            total_quantity = int(qty_price_m.group(1))
                            qty_conf = lines[j]['confidence']
                            j += 1

            # Format B: if no TU found, collect [GENCOD, size, qty(, price)] blocks
            if total_quantity == 0:
                qty_confs = []
                scan_j = gencod_pos  # restart from GENCOD position
                while scan_j < len(lines):
                    t = lines[scan_j]['text'].strip()
                    if re.match(r'^\d{13}$', t):
                        scan_j += 1
                        # Skip size line (digit or letter like S, M, L, XL)
                        if scan_j < len(lines):
                            size_t = lines[scan_j]['text'].strip()
                            if re.match(r'^(\d{1,2}|[SMLX]{1,3}|X?[SML]|[23]?XL)$', size_t, re.IGNORECASE):
                                scan_j += 1
                        # Qty + price: "qty price" on one line, or qty alone then price
                        if scan_j < len(lines):
                            qt = lines[scan_j]['text'].strip()
                            qty_price_m = re.match(r'^(\d+)\s+[\d,]+', qt)
                            if qty_price_m:
                                total_quantity += int(qty_price_m.group(1))
                                qty_confs.append(lines[scan_j]['confidence'])
                                scan_j += 1
                            elif re.match(r'^\d+$', qt):
                                # qty alone on this line, price on next
                                total_quantity += int(qt)
                                qty_confs.append(lines[scan_j]['confidence'])
                                scan_j += 1
                                # skip price line
                                if scan_j < len(lines) and re.match(r'^[\d,]+$', lines[scan_j]['text'].strip()):
                                    scan_j += 1
                            else:
                                scan_j += 1
                    elif t.upper().startswith('TOTAL') or re.match(r'^\d{6}\s+[A-Z]', t):
                        break
                    elif 'GIP' in t.upper() or 'SITG' in t.upper():
                        # GIP/SITG can appear mid-page, skip it
                        scan_j += 1
                    else:
                        scan_j += 1

                if total_quantity > 0:
                    j = scan_j
                    qty_conf = sum(qty_confs) / len(qty_confs) if qty_confs else 0.0

            if total_quantity > 0 and description:
                items.append({
                    'fournisseur': 'MULLIEZ-FLORY',
                    'date': date, 'date_conf': date_conf,
                    'reference': reference, 'ref_conf': ref_conf,
                    'article': description, 'article_conf': desc_conf,
                    'composants': None, 'composants_conf': 1.0,
                    'quantite': total_quantity, 'quantite_conf': qty_conf,
                })

            i = j if j > i + 1 else i + 1
        else:
            i += 1

    return items


def parse_poyet_motte(lines):
    """Parse Poyet Motte invoice lines.

    Two formats:
    - Direct format: "NEGOC-XXXXX DESCRIPTION", then "QTY UN ..." line
    - Chorus format: same as CTM (via parse_chorus_format)
    """
    # Try direct format first
    items = []
    date, date_conf = None, 0.0

    for line in lines:
        m = re.search(r'(\d{2}/\d{2}/\d{4})', line['text'])
        if m:
            try:
                dt = datetime.strptime(m.group(1), '%d/%m/%Y')
                if dt.year >= 2024:
                    t = line['text'].lower()
                    if 'ech' not in t and 'livraison' not in t and 'exp' not in t:
                        date = dt
                        date_conf = line['confidence']
                        break
            except ValueError:
                continue

    i = 0
    while i < len(lines):
        text = lines[i]['text'].strip()

        ref_m = re.match(r'^(NEGOC-\d+|[A-Z]{2,}\d*-\d+)\s+(.+)$', text)
        if ref_m:
            reference = ref_m.group(1)
            ref_conf = lines[i]['confidence']
            description_parts = [ref_m.group(2)]
            desc_conf_sum = lines[i]['confidence']
            desc_count = 1

            j = i + 1
            quantity = None
            qty_conf = 0.0

            while j < len(lines):
                t = lines[j]['text'].strip()
                qty_m = re.match(r'^(\d+)\s+UN\s+', t)
                if qty_m:
                    quantity = int(qty_m.group(1))
                    qty_conf = lines[j]['confidence']
                    j += 1
                    break
                if (not re.match(r'^[\d\s,.]+$', t) and
                    not re.match(r'^\d{13}$', t) and
                    'Port' not in t and 'Transporteur' not in t and
                    'Command' not in t and 'Livraison' not in t and
                    len(t) > 2):
                    description_parts.append(t)
                    desc_conf_sum += lines[j]['confidence']
                    desc_count += 1
                    j += 1
                else:
                    break

            description = ' '.join(description_parts)
            desc_conf = desc_conf_sum / desc_count

            if quantity is not None:
                items.append({
                    'fournisseur': 'POYET-MOTTE',
                    'date': date, 'date_conf': date_conf,
                    'reference': reference, 'ref_conf': ref_conf,
                    'article': description, 'article_conf': desc_conf,
                    'composants': None, 'composants_conf': 1.0,
                    'quantite': quantity, 'quantite_conf': qty_conf,
                })

            i = j
        else:
            i += 1

    # If direct format found nothing, try Chorus format
    if not items:
        items = parse_chorus_format(lines, 'POYET-MOTTE')

    return items


def parse_tissus_gisele(lines):
    """Parse Tissus Gisele invoice lines.

    Two formats:
    - Direct: "QTY,00 DESCRIPTION" where qty uses French number format
    - Chorus: same as CTM (via parse_chorus_format fallback)
    """
    items = []
    date, date_conf = None, 0.0

    # Try "le DD/MM/YYYY" first, then any DD/MM/YYYY
    for line in lines:
        m = re.search(r'le\s+(\d{2}/\d{2}/\d{4})', line['text'])
        if m:
            try:
                date = datetime.strptime(m.group(1), '%d/%m/%Y')
                date_conf = line['confidence']
                break
            except ValueError:
                continue

    i = 0
    while i < len(lines):
        text = lines[i]['text'].strip()

        m = re.match(r'^([\d.]+),00\s+(.+)$', text)
        if m:
            qty_str = m.group(1).replace('.', '')
            try:
                quantity = int(qty_str)
            except ValueError:
                i += 1
                continue

            if quantity < 1 or quantity > 100000:
                i += 1
                continue

            description = m.group(2).strip()

            # Skip false positives
            if len(description) < 3:
                i += 1
                continue
            # Skip Chorus-format lines ("1- ARTICLE" or "2-1 ARTICLE")
            if re.match(r'^\d+[\-.]', description):
                i += 1
                continue

            desc_conf = lines[i]['confidence']
            qty_conf = lines[i]['confidence']

            j = i + 1
            while j < len(lines):
                t = lines[j]['text'].strip()
                if re.match(r'^[\d.,\s]+$', t) and ',' in t:
                    break
                if t.lower().startswith('total') or t.lower().startswith('escompte'):
                    break
                if re.match(r'^[\d.]+,00\s+[A-Z]', t):
                    break
                if any(kw in t.lower() for kw in ['prix', 'page', 'suivant', 'net a payer']):
                    break
                description += ' ' + t
                desc_conf = (desc_conf + lines[j]['confidence']) / 2
                j += 1

            # Extract a reference code from the description
            reference = ''
            ref_conf = desc_conf
            code_m = re.search(r'([A-Z]{2,}\s*[A-Z0-9]*\s*\d+[Xx]\d+\s*[A-Z]*\s*[A-Z]*)', description)
            if code_m:
                reference = code_m.group(1).strip()
            else:
                words = description.split()[:4]
                reference = ' '.join(words)

            items.append({
                'fournisseur': 'TISSUS GISELE',
                'date': date, 'date_conf': date_conf,
                'reference': reference, 'ref_conf': ref_conf,
                'article': description, 'article_conf': desc_conf,
                'composants': None, 'composants_conf': 1.0,
                'quantite': quantity, 'quantite_conf': qty_conf,
            })

            i = j
        else:
            i += 1

    # If direct format found nothing, try Chorus format
    if not items:
        items = parse_chorus_format(lines, 'TISSUS GISELE')

    return items


# ── Excel Writer ─────────────────────────────────────────────────────

def write_excel(all_items, output_path, threshold=CONFIDENCE_THRESHOLD):
    """Write extracted items to Excel with yellow highlighting for low confidence."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '2025'

    headers = ['Fournisseur', 'Date', 'Référence Article', 'Article', 'Composants', 'Quantité', 'Fichier']
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    row_idx = 2

    items_by_supplier = {}
    for item in all_items:
        supplier = item['fournisseur']
        if supplier not in items_by_supplier:
            items_by_supplier[supplier] = []
        items_by_supplier[supplier].append(item)

    for supplier in SUPPLIER_ORDER:
        if supplier not in items_by_supplier:
            continue
        supplier_items = items_by_supplier[supplier]

        for idx, item in enumerate(supplier_items):
            if idx == 0:
                ws.cell(row=row_idx, column=1, value=supplier)

            cell_b = ws.cell(row=row_idx, column=2, value=item['date'])
            if item['date']:
                cell_b.number_format = 'DD/MM/YYYY'
            if item['date_conf'] < threshold:
                cell_b.fill = LOW_CONF_FILL

            cell_c = ws.cell(row=row_idx, column=3, value=item['reference'])
            if item['ref_conf'] < threshold:
                cell_c.fill = LOW_CONF_FILL

            cell_d = ws.cell(row=row_idx, column=4, value=item['article'])
            if item['article_conf'] < threshold:
                cell_d.fill = LOW_CONF_FILL

            cell_e = ws.cell(row=row_idx, column=5, value=item.get('composants'))
            if item.get('composants_conf', 1.0) < threshold:
                cell_e.fill = LOW_CONF_FILL

            cell_f = ws.cell(row=row_idx, column=6, value=item['quantite'])
            if item['quantite_conf'] < threshold:
                cell_f.fill = LOW_CONF_FILL

            ws.cell(row=row_idx, column=7, value=item.get('filename', ''))

            row_idx += 1

    total_row = row_idx
    ws.cell(row=total_row, column=1, value='Total')
    ws.cell(row=total_row, column=6, value=f'=SUM(F2:F{total_row - 1})')

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 40

    wb.save(output_path)
    print(f'Excel saved to {output_path} ({total_row - 2} items)')


# ── Main ─────────────────────────────────────────────────────────────

SUPPLIER_PARSERS = {
    "Cloro'fil Concept": parse_clorofil,
    "CTM STYLE": parse_ctm,
    "HALBOUT SAS": parse_halbout,
    "MULLIEZ-FLORY": parse_mulliez,
    "POYET-MOTTE": parse_poyet_motte,
    "TISSUS GISELE": parse_tissus_gisele,
}


def main():
    print('Loading OCR model...')
    model = ocr_predictor(pretrained=True)

    input_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), INPUT_DIR)
    pdf_files = sorted([f for f in os.listdir(input_dir) if f.lower().endswith('.pdf')])
    print(f'Found {len(pdf_files)} PDF files')

    all_items = []
    errors = []

    for idx, filename in enumerate(pdf_files, 1):
        supplier = detect_supplier(filename)
        if supplier is None:
            print(f'  [{idx}/{len(pdf_files)}] SKIP (unknown supplier): {filename}')
            continue

        filepath = os.path.join(input_dir, filename)
        print(f'  [{idx}/{len(pdf_files)}] Processing: {filename} -> {supplier}')

        try:
            export = ocr_pdf(filepath, model)
            lines = extract_lines(export)
            parser = SUPPLIER_PARSERS[supplier]
            items = parser(lines)

            if not items:
                print(f'    WARNING: no items extracted')
                errors.append((filename, 'no items extracted'))
            else:
                print(f'    Extracted {len(items)} items')
                for item in items:
                    item['filename'] = filename
                all_items.extend(items)

        except Exception as e:
            print(f'    ERROR: {e}')
            errors.append((filename, str(e)))

    print(f'\nTotal items extracted: {len(all_items)}')

    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), OUTPUT_FILE)
    write_excel(all_items, output_path, CONFIDENCE_THRESHOLD)

    print(f'\nConfidence threshold: {CONFIDENCE_THRESHOLD}')
    low_conf_count = sum(
        1 for item in all_items
        for conf in [item['date_conf'], item['ref_conf'], item['article_conf'], item['quantite_conf']]
        if conf < CONFIDENCE_THRESHOLD
    )
    print(f'Cells with low confidence: {low_conf_count}')

    if errors:
        print(f'\nErrors ({len(errors)}):')
        for fname, err in errors:
            print(f'  {fname}: {err}')


if __name__ == '__main__':
    main()
