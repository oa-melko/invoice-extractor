"""
Extract invoice data from scanned PDFs using docTR OCR + Ollama LLM and write to Excel.
Uses supplier-specific prompts for structured extraction.
"""
import os
import sys
import re
import json
from datetime import datetime
from doctr.io import DocumentFile
from doctr.models import ocr_predictor
import openpyxl
from openpyxl.styles import PatternFill, Font
import requests

# ── Configurable parameters ──────────────────────────────────────────
CONFIDENCE_THRESHOLD = 0.80
INPUT_DIR = "donnee a analyser"
OUTPUT_FILE = "Saisie Achat Linge - RESULTAT.xlsx"
OLLAMA_MODEL = "gemma2:latest"
OLLAMA_URL = "http://localhost:11434/api/generate"

# Supplier display names (order for Excel output)
SUPPLIER_ORDER = [
    "Cloro'fil Concept",
    "CTM STYLE",
    "HALBOUT SAS",
    "MULLIEZ-FLORY",
    "POYET-MOTTE",
    "TISSUS GISELE",
]

# Map filename prefixes to supplier keys
SUPPLIER_PREFIX_MAP = {
    "Clorofil": "clorofil",
    "CTM": "ctm_style",
    "Halbout": "halbout",
    "Mulliez": "mulliez_flory",
    "Poyet Motte": "poyet_motte",
    "Tissus Gis": "tissus_gisele",
}

# Map supplier keys to display names
SUPPLIER_DISPLAY = {
    "clorofil": "Cloro'fil Concept",
    "ctm_style": "CTM STYLE",
    "halbout": "HALBOUT SAS",
    "mulliez_flory": "MULLIEZ-FLORY",
    "poyet_motte": "POYET-MOTTE",
    "tissus_gisele": "TISSUS GISELE",
}

LOW_CONF_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# ── Supplier Prompts ─────────────────────────────────────────────────

SUPPLIER_PROMPTS = {
    "clorofil": """Rôle: Extracteur de facture textile professionnel
Tâche: Extraire les lignes articles de la facture Cloro'fil Concept

CHAMPS À EXTRAIRE:
1. Fournisseur: "Cloro'fil Concept"
2. Date: Date de facture (format: 05/09/2025)
3. Référence Article: Combiner les 2 lignes de référence
   - Ligne 1: Code numérique (ex: "000025")
   - Ligne 2: Code produit (ex: "GT350JNE")
   → Format final: "000025 GT350JNE"
4. Article: Première ligne de la désignation (ex: "GANT DE TOILETTE")
5. Composants: Spécifications techniques (ex: "350 g/m2 Jaune")
6. Quantité: Nombre entier de la colonne "Quantité" (ex: 13000)

RÈGLES SPÉCIFIQUES:
- La référence est toujours sur 2 niveaux dans le tableau
- Ignorer les lignes contenant "Prix au 100"
- Si plusieurs articles: créer une entrée par ligne article
- Le fournisseur peut être écrit "Cloro'fil Concept" ou "CLORO'FIL Concept"
""",

    "ctm_style": """Rôle: Extracteur facture électronique Chorus Pro
Tâche: Extraire les données de facture CTM Style (format Chorus)

CHAMPS À EXTRAIRE:
1. Fournisseur: "CTM STYLE"
2. Date: Date facture (ex: "14/11/2025")
3. Référence Article: "Référence produit" (EAN13 à 13 chiffres, ex: 3617540000462)
4. Article: "Dénomination de l'article" (ex: "ROMEO 80JF 313/40")
5. Composants: Laisser vide
6. Quantité: Colonne "Quantité facturée" (ex: 5,00 → 5)

RÈGLES SPÉCIFIQUES:
- Les articles sont dans la section "Articles rattachés au compte client"
- Un même article peut apparaître sur plusieurs lignes → créer une entrée par ligne
- Ignorer les lignes "Totaux du site de livraison" et "Brut HT/Net HT"
- Format des prix: virgule = séparateur décimal (format FR)
""",

    "halbout": """Rôle: Extracteur facture équipement hospitalier et textile
Tâche: Extraire les lignes articles de facture Halbout SAS

CHAMPS À EXTRAIRE:
1. Fournisseur: "HALBOUT SAS"
2. Date: Date facture (ex: "03/02/2025")
3. Référence Article: Colonne "Article" (code alphanumérique, ex: ORE06006025)
4. Article: Première partie de la désignation (nom du produit uniquement)
5. Composants: Description technique (dimensions, matière, normes)
6. Quantité: Colonne "Quantité" (entier, ex: 100)

RÈGLES SPÉCIFIQUES:
- Les descriptions sont très longues et techniques
- Séparer: nom produit → "Article", détails techniques → "Composants"
- Format: "OREILLER MICRONEW 60x60..." → Article: "OREILLER MICRONEW", Composants: "60x60 BLC IGNIFUGE..."
- Ignorer les lignes de livraison "Livré à:..."
- Ignorer les mentions BNP PARIBAS FACTOR et affacturage
""",

    "mulliez_flory": """Rôle: Extracteur facture textile professionnel (format industriel dense)
Tâche: Extraire les données du tableau complexe Mulliez-Flory

CHAMPS À EXTRAIRE:
1. Fournisseur: "MULLIEZ-FLORY"
2. Date: Date facture (ex: "04/06/2025")
3. Référence Article: Concaténer "REFERENCE" + "COLORIS" avec un tiret (ex: "036484-M01ZZ")
4. Article: Colonne "ARTICLE" (ex: "BAVOIR CONFORBEL CROCUS BP")
5. Composants: Colonne "COLORIS" + "TAILLE" (ex: "BEIGE MARRON 045*090")
6. Quantité: Colonne "QTES" - somme totale de toutes les tailles pour un même article

RÈGLES SPÉCIFIQUES:
- Un même article peut avoir plusieurs tailles avec des quantités différentes → SOMMER les quantités
- La référence est numérique (6 chiffres), le coloris est alphanumérique (4-5 caractères)
- NE PAS prendre la colonne "TOTAL" ni "TOTAL COMMANDE"
- Un même article peut avoir plusieurs coloris → Créer des lignes distinctes
""",

    "poyet_motte": """Rôle: Extracteur facture électronique Chorus Pro
Tâche: Extraire les données de facture Poyet Motte

CHAMPS À EXTRAIRE:
1. Fournisseur: "POYET-MOTTE"
2. Date: Date facture (ex: "07/11/2025")
3. Référence Article: "Référence produit" (EAN13, ex: 3120760082768)
4. Article: "Dénomination de l'article" (ex: "POLECO OUR SV")
5. Composants: Laisser vide
6. Quantité: "Quantité facturée" (ex: 300,00 → 300)

RÈGLES SPÉCIFIQUES:
- Même structure que CTM Style (Chorus Pro)
- Les articles sont dans la section "Articles rattachés au compte client"
- Ne pas confondre date de livraison avec date de facture
- Créer une entrée par ligne article
""",

    "tissus_gisele": """Rôle: Extracteur facture textile linge de maison
Tâche: Extraire les données de facture Tissus Gisèle

CHAMPS À EXTRAIRE:
1. Fournisseur: "TISSUS GISELE"
2. Date: Date facture (ex: "16/01/2025")
3. Référence Article: Colonne "Code" (ex: "DR C4 180X320 PC BLANCL2E")
4. Article: Première partie de la désignation (ex: "DRAPS PLATS OURLETS PIQUES DE 4CM AU FIL")
5. Composants: Suite de la désignation (dimensions, composition: "180,0 x 320,0 BLANC UNI POLYESTER-COTON 50/50")
6. Quantité: Colonne "Quantité" (format: 2.000,00 → convertir en 2000)

RÈGLES SPÉCIFIQUES:
- Format quantité: le point est un séparateur de milliers (2.000,00 = deux mille)
- Le code article contient des dimensions (DR C4 180X320...)
- "Commande SOLDEE" indique la fin du tableau
- Si format Chorus (avec "Référence produit"): utiliser la référence produit comme code article
""",
}

# ── OCR Functions ────────────────────────────────────────────────────

def detect_supplier(filename):
    """Detect supplier key from PDF filename prefix."""
    for prefix, key in SUPPLIER_PREFIX_MAP.items():
        if filename.startswith(prefix):
            return key
    return None


def ocr_pdf(filepath, model):
    """OCR a PDF file and return the docTR export dict."""
    doc = DocumentFile.from_pdf(filepath)
    result = model(doc)
    return result.export()


def extract_text_with_confidence(ocr_export):
    """Extract full text and per-line confidence from OCR export."""
    full_text = ""
    line_confidences = []
    for page in ocr_export['pages']:
        for block in page['blocks']:
            for line in block['lines']:
                text = ' '.join(w['value'] for w in line['words'])
                confs = [w['confidence'] for w in line['words']]
                avg = sum(confs) / len(confs) if confs else 0.0
                full_text += text + "\n"
                line_confidences.append(avg)
    avg_confidence = sum(line_confidences) / len(line_confidences) if line_confidences else 0.0
    return full_text, avg_confidence


# ── LLM Extraction ──────────────────────────────────────────────────

def call_ollama(prompt, ocr_text, model=OLLAMA_MODEL):
    """Send OCR text + prompt to Ollama and get structured JSON back."""
    # Truncate OCR text if too long to avoid context issues
    if len(ocr_text) > 6000:
        ocr_text = ocr_text[:6000]

    full_prompt = f"""{prompt}

TEXTE OCR DE LA FACTURE:
{ocr_text}

IMPORTANT: Réponds UNIQUEMENT avec un tableau JSON valide contenant les articles extraits.
Utilise ces clés exactes (sans accents): "fournisseur", "date", "reference_article", "article", "composants", "quantite"
Ne mets aucun texte avant ou après le JSON. Pas de commentaires. Juste le JSON.

Exemple de format:
[{{"fournisseur": "...", "date": "JJ/MM/AAAA", "reference_article": "...", "article": "...", "composants": "...", "quantite": 100}}]
"""

    try:
        response = requests.post(
            OLLAMA_URL,
            json={
                "model": model,
                "prompt": full_prompt,
                "stream": False,
                "options": {
                    "temperature": 0.1,
                    "num_predict": 4096,
                }
            },
            timeout=120,
        )
        response.raise_for_status()
        result = response.json()
        return result.get("response", "")
    except Exception as e:
        print(f"    Ollama error: {e}")
        return ""


def parse_llm_response(response_text):
    """Parse JSON from LLM response, handling common formatting issues."""
    text = response_text.strip()

    # Remove markdown code fences if present
    text = re.sub(r'^```json\s*', '', text)
    text = re.sub(r'^```\s*', '', text)
    text = re.sub(r'\s*```$', '', text)
    text = text.strip()

    # Try to find JSON array in the response
    match = re.search(r'\[.*\]', text, re.DOTALL)
    if match:
        text = match.group(0)

    try:
        data = json.loads(text)
        if isinstance(data, dict):
            data = [data]
        if isinstance(data, list):
            return data
    except json.JSONDecodeError:
        pass

    # Try to fix common issues (trailing commas, etc.)
    text = re.sub(r',\s*}', '}', text)
    text = re.sub(r',\s*]', ']', text)
    try:
        data = json.loads(text)
        if isinstance(data, dict):
            data = [data]
        return data if isinstance(data, list) else []
    except json.JSONDecodeError:
        return []


def extract_items_llm(ocr_text, supplier_key, filename):
    """Extract items from OCR text using LLM with supplier-specific prompt."""
    prompt = SUPPLIER_PROMPTS[supplier_key]
    supplier_name = SUPPLIER_DISPLAY[supplier_key]

    response = call_ollama(prompt, ocr_text)
    if not response:
        return []

    items = parse_llm_response(response)

    # Normalize and validate items
    valid_items = []
    for item in items:
        try:
            # Normalize keys (handle French accented keys from LLM)
            key_map = {
                'Fournisseur': 'fournisseur',
                'Date': 'date',
                'Référence Article': 'reference_article',
                'R\u00e9f\u00e9rence Article': 'reference_article',
                'Reference Article': 'reference_article',
                'reference': 'reference_article',
                'ref': 'reference_article',
                'Article': 'article',
                'Composants': 'composants',
                'Quantité': 'quantite',
                'Quantit\u00e9': 'quantite',
                'Quantite': 'quantite',
                'Fichier': 'filename',
            }
            normalized = {}
            for k, v in item.items():
                new_key = key_map.get(k, k.lower().replace(' ', '_'))
                normalized[new_key] = v
            item = normalized

            # Ensure correct supplier name
            item['fournisseur'] = supplier_name
            item['filename'] = filename

            # Normalize quantity
            qty = item.get('quantite', 0)
            if isinstance(qty, str):
                qty = qty.replace('.', '').replace(',', '').replace(' ', '')
                qty = int(re.sub(r'[^\d]', '', qty)) if qty else 0
            item['quantite'] = int(qty)

            # Normalize date
            date_str = item.get('date', '')
            item['date_parsed'] = None
            if date_str:
                for fmt in ['%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d']:
                    try:
                        item['date_parsed'] = datetime.strptime(date_str, fmt)
                        break
                    except ValueError:
                        continue

            # Ensure all fields exist
            item.setdefault('reference_article', '')
            item.setdefault('article', '')
            item.setdefault('composants', '')

            # Skip invalid items
            if item['quantite'] <= 0:
                continue
            if not item.get('reference_article') and not item.get('article'):
                continue

            valid_items.append(item)
        except (ValueError, TypeError, KeyError):
            continue

    return valid_items


# ── Excel Writer ─────────────────────────────────────────────────────

def write_excel(all_items, output_path, threshold=CONFIDENCE_THRESHOLD):
    """Write extracted items to Excel with yellow highlighting for low confidence."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '2025'

    headers = ['Fournisseur', 'Date', 'Référence Article', 'Article',
               'Composants', 'Quantité', 'Fichier']
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    row_idx = 2

    items_by_supplier = {}
    for item in all_items:
        supplier = item['fournisseur']
        if supplier not in items_by_supplier:
            items_by_supplier[supplier] = []
        items_by_supplier[supplier].append(item)

    for supplier in SUPPLIER_ORDER:
        if supplier not in items_by_supplier:
            continue
        supplier_items = items_by_supplier[supplier]

        for idx, item in enumerate(supplier_items):
            if idx == 0:
                ws.cell(row=row_idx, column=1, value=supplier)

            cell_b = ws.cell(row=row_idx, column=2, value=item.get('date_parsed'))
            if item.get('date_parsed'):
                cell_b.number_format = 'DD/MM/YYYY'
            if item.get('ocr_confidence', 1.0) < threshold:
                cell_b.fill = LOW_CONF_FILL

            cell_c = ws.cell(row=row_idx, column=3, value=item.get('reference_article', ''))
            if item.get('ocr_confidence', 1.0) < threshold:
                cell_c.fill = LOW_CONF_FILL

            cell_d = ws.cell(row=row_idx, column=4, value=item.get('article', ''))
            if item.get('ocr_confidence', 1.0) < threshold:
                cell_d.fill = LOW_CONF_FILL

            cell_e = ws.cell(row=row_idx, column=5, value=item.get('composants', ''))

            cell_f = ws.cell(row=row_idx, column=6, value=item['quantite'])
            if item.get('ocr_confidence', 1.0) < threshold:
                cell_f.fill = LOW_CONF_FILL

            ws.cell(row=row_idx, column=7, value=item.get('filename', ''))

            row_idx += 1

    total_row = row_idx
    ws.cell(row=total_row, column=1, value='Total')
    ws.cell(row=total_row, column=6, value=f'=SUM(F2:F{total_row - 1})')

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 55
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 40

    wb.save(output_path)
    print(f'Excel saved to {output_path} ({total_row - 2} items)')


# ── Main ─────────────────────────────────────────────────────────────

def main():
    input_dir = INPUT_DIR
    if len(sys.argv) > 1:
        input_dir = sys.argv[1]

    if not os.path.isabs(input_dir):
        input_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), input_dir)

    print('Loading OCR model...')
    ocr_model = ocr_predictor(pretrained=True)

    pdf_files = sorted([f for f in os.listdir(input_dir) if f.lower().endswith('.pdf')])
    print(f'Found {len(pdf_files)} PDF files')

    # Check Ollama is running
    try:
        r = requests.get("http://localhost:11434/api/tags", timeout=5)
        r.raise_for_status()
        print(f'Ollama connected (model: {OLLAMA_MODEL})')
    except Exception as e:
        print(f'ERROR: Cannot connect to Ollama: {e}')
        print('Make sure Ollama is running: ollama serve')
        return

    all_items = []
    errors = []

    for idx, filename in enumerate(pdf_files, 1):
        supplier_key = detect_supplier(filename)
        if supplier_key is None:
            print(f'  [{idx}/{len(pdf_files)}] SKIP (unknown supplier): {filename}')
            continue

        filepath = os.path.join(input_dir, filename)
        supplier_name = SUPPLIER_DISPLAY[supplier_key]
        print(f'  [{idx}/{len(pdf_files)}] Processing: {filename} -> {supplier_name}')

        try:
            # Step 1: OCR
            export = ocr_pdf(filepath, ocr_model)
            ocr_text, avg_confidence = extract_text_with_confidence(export)

            if not ocr_text.strip():
                print(f'    WARNING: empty OCR output')
                errors.append((filename, 'empty OCR'))
                continue

            # Step 2: LLM extraction
            items = extract_items_llm(ocr_text, supplier_key, filename)

            # Attach OCR confidence to each item
            for item in items:
                item['ocr_confidence'] = avg_confidence

            if not items:
                print(f'    WARNING: no items extracted by LLM')
                errors.append((filename, 'no items from LLM'))
            else:
                print(f'    Extracted {len(items)} items (OCR conf: {avg_confidence:.2f})')
                all_items.extend(items)

        except Exception as e:
            print(f'    ERROR: {e}')
            errors.append((filename, str(e)))

    print(f'\nTotal items extracted: {len(all_items)}')

    # Write Excel
    output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), OUTPUT_FILE)
    write_excel(all_items, output_path, CONFIDENCE_THRESHOLD)

    # Summary
    print(f'\nConfidence threshold: {CONFIDENCE_THRESHOLD}')
    low_conf = sum(1 for i in all_items if i.get('ocr_confidence', 1.0) < CONFIDENCE_THRESHOLD)
    print(f'Items with low OCR confidence: {low_conf}')

    if errors:
        print(f'\nErrors ({len(errors)}):')
        for fname, err in errors:
            print(f'  {fname}: {err}')


if __name__ == '__main__':
    main()
